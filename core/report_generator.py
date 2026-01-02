"""
Report Generation Module
Generates comprehensive fiscal policy reports in PDF and Excel formats.

Features:
- Export policy analysis to PDF with charts
- Excel workbooks with multiple sheets
- Summary statistics and tables
- Customizable report sections
"""

import os
import io
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
import pandas as pd
import numpy as np

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    OPENPYXL_ERROR = str(e)


@dataclass
class ReportSection:
    """A section of the report."""
    title: str
    content: str
    data: Optional[pd.DataFrame] = None
    chart_path: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ReportMetadata:
    """Metadata for the report."""
    title: str
    author: str = "PoliSim"
    date: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))
    organization: str = "Galactic Office of Development"
    version: str = "1.0"
    description: str = ""


class PDFReportGenerator:
    """Generate PDF reports from policy analysis."""
    
    def __init__(self, metadata: Optional[ReportMetadata] = None):
        """Initialize PDF generator."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab required for PDF generation. Install with: pip install reportlab")
        
        self.metadata = metadata or ReportMetadata(title="Policy Analysis Report")
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles."""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER,
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2e5c8a'),
            spaceAfter=12,
            spaceBefore=12,
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['BodyText'],
            fontSize=10,
            alignment=TA_JUSTIFY,
            spaceAfter=8,
        ))
    
    def generate(
        self,
        sections: List[ReportSection],
        output_path: str,
        include_toc: bool = True,
    ) -> str:
        """
        Generate a PDF report.
        
        Args:
            sections: List of report sections
            output_path: Path to save PDF
            include_toc: Include table of contents
        
        Returns:
            Path to generated PDF
        """
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        
        # Title page
        story.append(Paragraph(self.metadata.title, self.styles['CustomTitle']))
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(
            f"<b>Date:</b> {self.metadata.date}<br/>"
            f"<b>Author:</b> {self.metadata.author}<br/>"
            f"<b>Organization:</b> {self.metadata.organization}",
            self.styles['Normal']
        ))
        
        if self.metadata.description:
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph(self.metadata.description, self.styles['CustomBody']))
        
        story.append(PageBreak())
        
        # Table of contents
        if include_toc and sections:
            story.append(Paragraph("Table of Contents", self.styles['CustomHeading']))
            for i, section in enumerate(sections, 1):
                story.append(Paragraph(f"{i}. {section.title}", self.styles['Normal']))
            story.append(PageBreak())
        
        # Sections
        for section in sections:
            story.append(Paragraph(section.title, self.styles['CustomHeading']))
            
            if section.content:
                story.append(Paragraph(section.content, self.styles['CustomBody']))
                story.append(Spacer(1, 0.1*inch))
            
            if section.data is not None:
                table_data = self._dataframe_to_table_data(section.data)
                table = self._create_styled_table(table_data)
                story.append(table)
                story.append(Spacer(1, 0.2*inch))
            
            if section.chart_path and os.path.exists(section.chart_path):
                try:
                    img = Image(section.chart_path, width=6*inch, height=3.5*inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2*inch))
                except Exception as e:
                    print(f"Warning: Could not embed chart {section.chart_path}: {e}")
            
            story.append(PageBreak())
        
        # Generate PDF
        doc.build(story)
        return output_path
    
    def _dataframe_to_table_data(self, df: pd.DataFrame, max_rows: int = 50) -> List[List[str]]:
        """Convert DataFrame to table data for ReportLab."""
        if len(df) > max_rows:
            df = df.head(max_rows)
        
        # Header row
        table_data = [list(df.columns.astype(str))]
        
        # Data rows
        for _, row in df.iterrows():
            table_data.append([str(val) for val in row.values])
        
        return table_data
    
    def _create_styled_table(self, data: List[List[str]]) -> "Table":
        """Create a styled ReportLab table."""
        table = Table(data, colWidths=[1.5*inch] * len(data[0]))
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e5c8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        return table


class ExcelReportGenerator:
    """Generate Excel reports from policy analysis."""
    
    def __init__(self, metadata: Optional[ReportMetadata] = None):
        """Initialize Excel generator."""
        if not OPENPYXL_AVAILABLE:
            error_msg = f"openpyxl required for Excel generation. Install with: pip install openpyxl\nError: {OPENPYXL_ERROR if 'OPENPYXL_ERROR' in globals() else 'Unknown import error'}"
            raise ImportError(error_msg)
        
        self.metadata = metadata or ReportMetadata(title="Policy Analysis Report")
    
    def generate(
        self,
        sections: Dict[str, pd.DataFrame],
        summary: Optional[Dict[str, Any]] = None,
        output_path: str = "report.xlsx",
    ) -> str:
        """
        Generate an Excel workbook report.
        
        Args:
            sections: Dictionary of sheet_name -> DataFrame
            summary: Summary statistics dictionary
            output_path: Path to save Excel file
        
        Returns:
            Path to generated Excel file
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            
            # Summary sheet
            row = 1
            try:
                ws[f'A{row}'] = self.metadata.title
                self._style_header_cell(ws[f'A{row}'])
            except (TypeError, AttributeError, ValueError) as e:
                # Skip styling if cell access fails
                pass
            row += 1
            
            try:
                ws[f'A{row}'] = f"Generated: {self.metadata.date}"
            except (TypeError, AttributeError, ValueError):
                pass
            row += 1
            
            try:
                ws[f'A{row}'] = f"Author: {self.metadata.author}"
            except (TypeError, AttributeError, ValueError):
                pass
            row += 2
            
            if summary:
                try:
                    ws[f'A{row}'] = "Summary Statistics"
                    self._style_subheader_cell(ws[f'A{row}'])
                except (TypeError, AttributeError, ValueError):
                    pass
                row += 1
                
                for key, value in summary.items():
                    try:
                        ws[f'A{row}'] = str(key)
                        ws[f'B{row}'] = value
                    except (TypeError, AttributeError, ValueError):
                        pass
                    row += 1
            
            # Data sheets
            for sheet_name, df in sections.items():
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                
                # Skip empty dataframes
                if df is None or (isinstance(df, pd.DataFrame) and df.empty):
                    continue
                
                ws = wb.create_sheet(sheet_name)
                self._write_dataframe_to_sheet(ws, df)
            
            wb.save(output_path)
            return output_path
        except KeyError as ke:
            # Re-raise with more context for debugging
            raise RuntimeError(f"KeyError in Excel generation: '{ke}' - this usually means a column name is invalid or a dictionary key doesn't exist") from ke
        except Exception as e:
            # Re-raise with more context
            raise RuntimeError(f"Excel generation failed: {type(e).__name__}: {str(e)}") from e
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """Write DataFrame to worksheet with formatting."""
        try:
            # Headers
            for col_idx, col_name in enumerate(df.columns, 1):
                try:
                    cell = ws.cell(row=1, column=col_idx, value=str(col_name))
                    self._style_header_cell(cell)
                except Exception:
                    # Skip header if it fails
                    pass
            
            # Data - with robust type conversion
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row.values, 1):
                    try:
                        # Convert problematic types before writing
                        cell_value = self._sanitize_cell_value(value)
                        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                        self._style_data_cell(cell)
                    except Exception as cell_error:
                        # If even sanitized value fails, skip this cell
                        pass
            
            # Adjust column widths
            for col_idx in range(1, len(df.columns) + 1):
                max_length = 0
                column_letter = self._get_column_letter(col_idx)
                
                for row_idx in range(1, len(df) + 2):
                    try:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                
                try:
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                except Exception:
                    pass
        except Exception as e:
            # Log and re-raise with context
            import traceback
            error_msg = f"Error writing dataframe to sheet: {type(e).__name__}: {str(e)}\n{traceback.format_exc()}"
            raise RuntimeError(error_msg) from e
    
    def _sanitize_cell_value(self, value):
        """Convert any value to something Excel can handle."""
        import numpy as np
        
        # Handle None and NaN
        if value is None:
            return None
        if pd.isna(value):
            return None
        
        # Handle numpy types
        if isinstance(value, np.integer):
            return int(value)
        if isinstance(value, np.floating):
            if np.isnan(value) or np.isinf(value):
                return None
            return float(value)
        
        # Handle Python numbers
        if isinstance(value, (int, float)):
            if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
                return None
            return value
        
        # Handle booleans
        if isinstance(value, (bool, np.bool_)):
            return bool(value)
        
        # Handle datetime
        if isinstance(value, (pd.Timestamp, pd.datetime)):
            return value
        
        # For everything else, convert to string (safest fallback)
        try:
            # Try to convert to string, but check if it looks like a Range object
            str_value = str(value)
            if 'Range' in str_value or 'Cell' in str_value:
                # Looks like an openpyxl object, skip it
                return None
            return str_value
        except Exception:
            # Last resort: return None
            return None
    
    def _style_header_cell(self, cell):
        """Style a header cell."""
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
    
    def _style_subheader_cell(self, cell):
        """Style a subheader cell."""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4A7BA7", end_color="4A7BA7", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def _style_data_cell(self, cell):
        """Style a data cell."""
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color="D3D3D3"),
            right=Side(style='thin', color="D3D3D3"),
            top=Side(style='thin', color="D3D3D3"),
            bottom=Side(style='thin', color="D3D3D3"),
        )
    
    @staticmethod
    def _get_column_letter(col_idx: int) -> str:
        """Convert column number to letter."""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(65 + col_idx % 26) + result
            col_idx //= 26
        return result


class ComprehensiveReportBuilder:
    """Build comprehensive policy analysis reports."""
    
    def __init__(self, metadata: Optional[ReportMetadata] = None):
        """Initialize report builder."""
        self.metadata = metadata or ReportMetadata(title="Policy Analysis Report")
        self.sections: List[ReportSection] = []
        self.summary_data: Dict[str, Any] = {}
    
    def add_executive_summary(self, summary_text: str) -> "ComprehensiveReportBuilder":
        """Add executive summary section."""
        self.sections.append(ReportSection(
            title="Executive Summary",
            content=summary_text,
        ))
        return self
    
    def add_policy_overview(
        self,
        policy_name: str,
        revenue_impact: float,
        spending_impact: float,
        deficit_impact: float,
    ) -> "ComprehensiveReportBuilder":
        """Add policy overview section."""
        content = (
            f"<b>{policy_name}</b><br/><br/>"
            f"<b>Revenue Impact:</b> ${revenue_impact:+,.0f}B<br/>"
            f"<b>Spending Impact:</b> ${spending_impact:+,.0f}B<br/>"
            f"<b>Deficit Impact:</b> ${deficit_impact:+,.0f}B<br/>"
        )
        
        self.sections.append(ReportSection(
            title="Policy Overview",
            content=content,
        ))
        
        self.summary_data.update({
            "Revenue Impact (B)": revenue_impact,
            "Spending Impact (B)": spending_impact,
            "Deficit Impact (B)": deficit_impact,
        })
        
        return self
    
    def add_fiscal_projections(self, projections_df: pd.DataFrame) -> "ComprehensiveReportBuilder":
        """Add fiscal projections table."""
        self.sections.append(ReportSection(
            title="10-Year Fiscal Projections",
            content="Year-by-year projections of revenue, spending, and deficit under the proposed policy.",
            data=projections_df,
        ))
        return self
    
    def add_sensitivity_analysis(self, sensitivity_df: pd.DataFrame) -> "ComprehensiveReportBuilder":
        """Add sensitivity analysis."""
        self.sections.append(ReportSection(
            title="Sensitivity Analysis",
            content="Impact of parameter variations on deficit outcomes.",
            data=sensitivity_df,
        ))
        return self
    
    def add_scenario_comparison(self, scenarios_df: pd.DataFrame) -> "ComprehensiveReportBuilder":
        """Add scenario comparison."""
        self.sections.append(ReportSection(
            title="Scenario Comparison",
            content="Comparison of multiple policy scenarios.",
            data=scenarios_df,
        ))
        return self
    
    def add_monte_carlo_results(self, results_df: pd.DataFrame) -> "ComprehensiveReportBuilder":
        """Add Monte Carlo simulation results."""
        self.sections.append(ReportSection(
            title="Monte Carlo Analysis",
            content=(
                "Stochastic simulation results showing confidence bounds and "
                "probability distributions of fiscal outcomes."
            ),
            data=results_df,
        ))
        return self
    
    def add_recommendations(self, recommendations_text: str) -> "ComprehensiveReportBuilder":
        """Add recommendations section."""
        self.sections.append(ReportSection(
            title="Recommendations",
            content=recommendations_text,
        ))
        return self
    
    def add_technical_appendix(self, appendix_df: pd.DataFrame, title: str = "Technical Details") -> "ComprehensiveReportBuilder":
        """Add technical appendix."""
        self.sections.append(ReportSection(
            title=title,
            content="Technical assumptions and methodology details.",
            data=appendix_df,
        ))
        return self
    
    def generate_pdf(self, output_path: str) -> str:
        """Generate PDF report."""
        generator = PDFReportGenerator(self.metadata)
        return generator.generate(self.sections, output_path)
    
    def generate_excel(self, output_path: str) -> str:
        """Generate Excel report."""
        # Convert sections to dictionary of DataFrames
        sheets = {}
        for section in self.sections:
            if section.data is not None:
                sheets[section.title] = section.data
        
        generator = ExcelReportGenerator(self.metadata)
        return generator.generate(sheets, self.summary_data, output_path)
    
    def generate_json(self, output_path: str) -> str:
        """Generate JSON report."""
        report_data = {
            "metadata": {
                "title": self.metadata.title,
                "date": self.metadata.date,
                "author": self.metadata.author,
            },
            "sections": []
        }
        
        for section in self.sections:
            section_data = {
                "title": section.title,
                "content": section.content,
            }
            
            if section.data is not None:
                section_data["data"] = section.data.to_dict(orient="records")
            
            report_data["sections"].append(section_data)
        
        with open(output_path, 'w') as f:
            json.dump(report_data, f, indent=2, default=str)
        
        return output_path

    def generate_html(self, output_path: str) -> str:
        """Generate a simple HTML report (no external deps)."""
        parts: List[str] = []
        parts.append("<html><head><meta charset='utf-8'><title>{}</title>".format(self.metadata.title))
        parts.append("<style>body{font-family:Arial, sans-serif; margin:24px;} h1,h2{color:#1f4788;} table{border-collapse:collapse; width:100%; margin-bottom:16px;} th,td{border:1px solid #ccc; padding:6px;} th{background:#2e5c8a; color:#fff;}</style>")
        parts.append("</head><body>")
        parts.append(f"<h1>{self.metadata.title}</h1>")
        parts.append(f"<p><strong>Date:</strong> {self.metadata.date}<br/><strong>Author:</strong> {self.metadata.author}<br/><strong>Description:</strong> {self.metadata.description}</p>")

        for section in self.sections:
            parts.append(f"<h2>{section.title}</h2>")
            if section.content:
                parts.append(f"<p>{section.content}</p>")
            if section.data is not None and isinstance(section.data, pd.DataFrame):
                parts.append(section.data.to_html(index=False, border=0))

        parts.append("</body></html>")

        Path(output_path).write_text("\n".join(parts), encoding="utf-8")
        return output_path
